"""
Created on Tue Jun  2 13:36:55 2020

@author: suraj

Custom environment to optimize the shape of an airfoil using OpenFoam CFD solver

"""

import gym
import os
import csv
import pandas as pd
from openpyxl import load_workbook
from gym import error, spaces, utils
from gym.utils import seeding
import numpy as np
import matplotlib.pyplot as plt
from scipy.integrate import odeint
from mpl_toolkits.mplot3d import Axes3D

from parametric_airfoil import *

from os import path
from PyFoam.RunDictionary.SolutionDirectory import SolutionDirectory
from PyFoam.RunDictionary.ParsedParameterFile import ParsedParameterFile
from PyFoam.Basics.DataStructures import Vector
from PyFoam.Execution.UtilityRunner import UtilityRunner
from PyFoam.Execution.BasicRunner import BasicRunner
from PyFoam.Infrastructure.ClusterJob import SolverJob
from PyFoam.RunDictionary.SolutionFile import SolutionFile
from PyFoam.RunDictionary.SolutionDirectory import SolutionDirectory
from PyFoam.RunDictionary.ParsedParameterFile import ParsedParameterFile
from PyFoam.Execution.ParallelExecution import LAMMachine
from PyFoam.Basics.DataStructures import Vector
from PyFoam.Error import error

import pyvista as vtki

import time
from time import gmtime, strftime
import logging
import warnings

import sys
import subprocess
from subprocess import Popen, PIPE, CalledProcessError

# global variables related to OpenFoam

mesh = 'blockMesh'              # command in OpenFoam to perform the meshing
solver = 'simpleFoam'           # command in OpenFoam to solve NS equations
solveroptions = '-case'         # solver options
origcase = 'baseCase'           # directory path of the base case file

#logger_g = logging.getLogger(__name__)
#logging.basicConfig(
#        filename='foamstatus.log',
#        format='%(asctime)s | %(message)s',
#        datefmt='%m/%d/%Y %I:%M:%S %p',
#        level=logging.INFO)

#%%
class shape_optimization(gym.Env):
    '''
    This class is a fym environment class for shape-optimization of an airfoil. The gym environment
    should have follwing functions:
        
        __init__ : initializes the class and some variables
        reset    : this function assigns the initial state of the environment
        step     : this function performs an action in the environment and updates the state of the
                   environment and corresponding reward of the action
    
    '''
    metadata = {'render.modes': ['human'],
                'video.frames_per_second' : 30}
    
    # Initialize the parameters for the Airfoil shape optimization and the RL
    def __init__(self,env_config):
        
        # if the baseCase folder has some logs from older simulation, remove them
        with open(f'{origcase}/logr.remove', 'wb') as fp:
            subprocess.run(
                f'rm {origcase}/log.* {origcase}/*.txt',
                shell=True,
                stdout=fp,
                stderr=subprocess.STDOUT
            )
            
        # if the baseCase folder has some solution from older simulation, remove them
        # OpenFoam saves the results in a folder which has the name = iteration number
        # 0 folder will have initial solution, 100 folder will have solution at 100th iteration and so on
        with open(f'{origcase}/logr.remove', 'a') as fp:
            subprocess.run(
                f'rm -r {origcase}/0.* {origcase}/[1-9]*',
                shell=True,
                stdout=fp,
                stderr=subprocess.STDOUT
            )
        
        # foamToVTK [options] command is used in OpenFoam to conver the data into vtk formart
        # it is easy to deal with the data saved in vtk format and there are some good libraries
        # to extract and modify data 
        # Good resource: http://www.shaowupan.com/posts/2018/09/blog-post-modify-vtk-openfoam/
        
        # if the baseCase folder has some solution from older simulation saved in VTK folder, remove them
        with open(f'{origcase}/logr.remove', 'a') as fp:
            subprocess.run(
                f'rm -r {origcase}/VTK',
                shell=True,
                stdout=fp,
                stderr=subprocess.STDOUT
            )
        
        # if there are additional folders like postProcessing, remove them
#        subprocess.run(
#            f'rm -r postProcessing',
#            shell=True,
#            stderr=subprocess.STDOUT
#        )
        
        # if running this file to test the class
        if env_config['test']:
            # for testing worker_index = 1
            self.worker_index = env_config['worker_index']
        else:
            # when actually running the RL algorithm for distributed training, the processor ID (PID)
            # is assiggned as the worker_index
            # in RlLib the PID can be accessed by worker_index of env_config
            self.worker_index = env_config.worker_index
            
        self.number_steps = 0
        self.end_time = env_config['end_time']
        self.write_interval = env_config['write_interval']
        
        self.vx = env_config['vx']
        self.control_points_low = env_config['controlparams_low']
        self.control_points_high = env_config['controlparams_high']
        
        self.res_ux_tol = env_config['res_ux_tol']
        self.res_uy_tol = env_config['res_uy_tol']
        self.res_p_tol = env_config['res_p_tol']
        self.res_nutilda_tol = env_config['res_nutilda_tol']
        
        self.reward_type = env_config['reward_type']
        self.states_type = env_config['states_type']
        
        num_panels = 50
        num_pts = num_panels + 1
        
        x_rad = np.linspace(0, np.pi, num_pts)
        self.x_cos = (np.cos(x_rad) / 2) + 0.5
        self.x_cos = self.x_cos[1:]
        self.num_cp = 3
        
        self.casename = 'baseCase_'+str(self.worker_index)
        self.csvfile = 'progress_'+str(self.worker_index)+'.csv'  

        self.state = None
        
        # Define the bounded action space
        # refer to https://gym.openai.com/docs/ for more details on spaces
        self.action_space = spaces.Box(self.control_points_low, self.control_points_high, dtype=np.float64)
        
        # Define the unbounded state-space
        if self.states_type == 1:
            high1 = np.array([np.finfo(np.float64).max])
        elif self.states_type == 2:
            high1 = np.array([np.finfo(np.float64).max,np.finfo(np.float64).max,np.finfo(np.float64).max,
                          np.finfo(np.float64).max,np.finfo(np.float64).max])
    
        self.observation_space = spaces.Box(-high1, high1, dtype=np.float64)
        
        # penalty for divergence (a hyperprameter)
        self.high_penalty = 10
        
        self.seed()
            
    # Seed for random number generator
    def seed(self, seed=None):
        self.np_random, seed = seeding.np_random(seed)
        return [seed]

    def reset(self):   
        '''
        This function resets the environment. In distributed RL training, the environment will be
        evaluated simultaneosuly on different processors. Each processor will evaluate multiple 
        episodes (i.e., one episode is a single CFD simulation for this problem)
        Tis function initilizaes the OpenFoam baseCase corresponding to a specific processor
        '''
        
        pid = self.worker_index 
#        logger_g.info(f'{pid}')
        
        # make a replica of the baseCase using PyFoam
        self.casename = 'baseCase_'+str(pid)
        self.csvfile = 'progress_'+str(pid)+'.csv'  
        orig = SolutionDirectory(origcase,archive=None,paraviewLink=False)
        case = orig.cloneCase(self.casename )
        
        # PyFoam library is a library that will allow the user to modify the parameters
        # of the CFD simulations in a user-friendly manner
        control_dict = ParsedParameterFile(path.join(self.casename,"system", "controlDict"))
        control_dict["endTime"] = self.end_time
        control_dict["writeInterval"] = self.write_interval
        control_dict.writeFile()
        
        # copy the blockMeshDictGenerator.xlsx file from baseCase to baseCase folder corresponding 
        # a specific processor
        # blockMeshDictGenerator.xlsx file has all formulas that uses and airfoil shape to 
        # generate a mesh around an airfoil
        # taken from : https://www.phanquocthien.org/mesh-geometry/blockmesh/airfoil
        with open(f'log.convert', 'a') as fp:
            subprocess.run(
                f'cp blockMeshDictGenerator.xlsx ./{self.casename}/blockMeshDictGenerator.xlsx',
                shell=True,
                stdout=fp,
                stderr=subprocess.STDOUT
                ) 
            
        # the baseCase_{pid} folder will be updated for every episode of the RL as new CFD simulation
        # will be performed fro each episode. Therefore, this folder will have some solution 
        # from older simulation that needs to be remove at the start of new episode
        with open(f'{self.casename}/logr.remove', 'a') as fp:
            subprocess.run(
                f'rm {self.casename}/log.*',
                shell=True,
                stdout=fp,
                stderr=subprocess.STDOUT
                )
            
        # remove old solution directories
        with open(f'{self.casename}/logr.remove', 'a') as fp:
            subprocess.run(
                f'rm -r {self.casename}/0.* {self.casename}/[1-9]*',
                shell=True,
                stdout=fp,
                stderr=subprocess.STDOUT
            )
        
        # remove old solution saved in the VTK format
        with open(f'{self.casename}/logr.remove', 'a') as fp:
            subprocess.run(
                f'rm -r {self.casename}/VTK',
                shell=True,
                stdout=fp,
                stderr=subprocess.STDOUT
            )
        
        # remove additional directories
#        subprocess.run(
#            f'rm -r postProcessing',
#            shell=True,
#            stderr=subprocess.STDOUT
#        )
        
        # assign the state
        if self.states_type == 1:
            # inlet velocty as the state of the RL algorithm
            self.state = np.array([self.vx]) 
        elif self.states_type == 2:
            # no state 
            None
                    
        return np.array(self.state)
    
    # Update the state of the environment
    def step(self, action):
        '''
        This function performas an action on the environemnt and then updates the state of the system. 
        The function returns the new state of the sytem, the reward of the action, and a flag to 
        indicate whether the episode is finished or not.
        For the shape optimization problem, one step of the RL is a single episode. Therefore the 
        episode will finish ath evry step of the RL.
        '''
        
        # check if the action is within lower and upper bound
        assert self.action_space.contains(action) , "%r (%s) invalid" %(action, type(action))
        
        pid = self.worker_index #os.getpid()
        logger_m = logging.getLogger(__name__)
        logging.basicConfig(filename=f'foamstatus_{pid}.log',
                            format='%(asctime)s | %(message)s',
                            datefmt='%m/%d/%Y %I:%M:%S %p',
                            level=logging.INFO)
        
        # episode is not finished
        done = False
        
        # assign action to control points
        action_cp = action
        
        # generate the airfoil shape using control points
        # there are different ways in the airfoil shape can be parameterized
        # I used a Github code to do this
        # GirHub repo code: 
        #https://github.com/nathanrooy/aerodynamic-shape-optimization/blob/master/pyfoil/parametric_airfoil.py
        
        airfoil_shape = np.array(bezier_airfoil(self.x_cos, munge_ctlpts(action_cp, self.num_cp, self.num_cp)))
        
        # import the workbook from blockMeshDictGenerator.xlsx
        wb = load_workbook(f'./{self.casename}/blockMeshDictGenerator.xlsx')
        ws = wb['Input'] 
        shift = 9
        
        # copies the airfoil shape in Input sheet of the blockMeshDictGenerator.xlsx file
        for index, data in enumerate(airfoil_shape):
            ws.cell(row=index+shift, column=1).value = data[0]
            ws.cell(row=index+shift, column=2).value = data[1]
        
        # for some reason, I had hard time in generating blockMeshDict in OpenFoam format from 
        # blockMeshDictGenerator.xlsx file
        # I had to go from .xlsx --> .xls --> .csv and then make some changes in .csv file
        # and then save blockMeshDict in OpenFoam format
        wb.save(f'./{self.casename}/blockMeshDictGenerator.xls')
        
        with open(f'log.convert', 'a') as fp:
            subprocess.run(
                    f'libreoffice --headless --convert-to xlsx ./{self.casename}/*.xls --outdir ./{self.casename}',
                    shell=True,
                    stdout=fp,
                    stderr=subprocess.STDOUT
                    )    
        
        wb = load_workbook(f'./{self.casename}/blockMeshDictGenerator.xlsx',data_only=True)
        sh = wb['blockMesh'] 
        
        with open(f'./{self.casename}/blockMeshDictGenerator.csv', 'w') as f:  
            c = csv.writer(f)
            for r in sh.rows:
                c.writerow([cell.value for cell in r])
                
        pd.set_option("display.max_colwidth", 100)
        df = pd.read_csv(f'./{self.casename}/blockMeshDictGenerator.csv',header=None)
        df = df.fillna('')
        
        # quick fix
        for i in [8,11,12]:
            for j in [47,82,102,142]:
                df[i][j] = str(int(df[i][j]))
                    
        base_filename = f'./{self.casename}/system/blockMeshDict'

        df.to_csv(base_filename, header=None, index=None, quoting=csv.QUOTE_NONE, sep=' ', escapechar = ' ')
        
        now = strftime("%m.%d.%Y-%H.%M.%S", gmtime())
        meshLogFile= f'log.{mesh}-{now}'
        
        # generate a mesh for an airfoil       
        proc = subprocess.Popen([f'$FOAM_APPBIN/{mesh} {solveroptions} {self.casename} >> {self.casename}/{meshLogFile}'],
                        shell=True,
                        stdout=subprocess.PIPE, 
                        stderr=subprocess.PIPE)
        
        proc.wait()
        (stdout, stderr) = proc.communicate()
        
        if proc.returncode != 0:
            # if the mesh fails, write the information in logger (for debugging purposes)
            logger_m.info(f'{action_cp[0]} mesh failed on {pid}')
        else:
            # mesh successful, run the simulation
            now = strftime("%m.%d.%Y-%H.%M.%S", gmtime())
            solverLogFile= f'log.{solver}-{now}'
            
            # to run on theta       
            proc = subprocess.Popen([f'$FOAM_APPBIN/{solver} {solveroptions} {self.casename} >> {self.casename}/{solverLogFile}'],
                            shell=True,
                            stdout=subprocess.PIPE, 
                            stderr=subprocess.PIPE)
                    
            proc.wait()
            (stdout, stderr) = proc.communicate()
        
        if proc.returncode != 0:
            # if the solution diverges, assign a negative reward (penalize that action)
            # episode is finished
            done = True
            reward = -self.high_penalty
            res_ux,res_uy,res_p,res_nutilda = 1.0, 1.0, 1.0, 1.0
            logger_m.info(f'{action_cp[0]} coefficient solution diverged in the first step on {pid}')
        else:
            # if the CFD simulation is converges, extract residuals
            # extract_residual.sh is a simple bash script to extract residual from OpenFoam
            # there are much more efficient and clener ways to do it
            with open(f'{self.casename}/logr.foamLog', 'a') as fp:   
                subprocess.run(
                        f'./extract_residual.sh {self.casename}/{solverLogFile} {self.casename} {pid}',
                        shell=True,
                        stdout = fp,
                        stderr=subprocess.STDOUT
                        )
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    log_history = np.genfromtxt(f'./{self.casename}/residual_{pid}.txt')
                
                # if log_histyory is empty, it means that something broke with either CFD simulation
                # or the residuals are too large. 
                # penalize that action
                if log_history.size == 0:
                    done = True
                    reward = -self.high_penalty
                    res_ux,res_uy,res_p,res_nutilda = 1.0, 1.0, 1.0, 1.0
                    logger_m.info(f'{action_cp[0]} coefficient solution no residual history')
                else:
                    # solution is converged, assign the residuals
                    log_history = np.reshape(log_history,[-1,5])
                    itercount = int(log_history[-1,0])
                    res_ux = log_history[-1,1]
                    res_uy = log_history[-1,2]
                    res_p = log_history[-1,3]
                    res_nutilda = log_history[-1,4]

        # convert the solution in vtk format
        proc = subprocess.Popen([f'$FOAM_APPBIN/foamToVTK {solveroptions} {self.casename} >> {self.casename}/logr.vtkoutput'],
                            shell=True,
                            stdout=subprocess.PIPE, 
                            stderr=subprocess.PIPE)

        proc.wait()
        (stdout, stderr) = proc.communicate()
        
        # solution has blown up, it gets caught while converting the solution to VTK format
        # penalize that action
        if proc.returncode != 0:
            done = True
            reward = -self.high_penalty
            res_ux,res_uy,res_p,res_nutilda = 1.0, 1.0, 1.0, 1.0
            logger_m.info(f'{action_cp[0]} coefficient solution diverged with large magnitude of velocity on {pid}')
        
        # if everything is successful, compute the lift and drag coefficient
        if (res_ux < self.res_ux_tol and res_uy < self.res_uy_tol and res_p < self.res_p_tol and \
            res_nutilda < self.res_nutilda_tol ) or itercount == self.end_time:    
            
            logger_m.info(f'{action_cp[0]} coefficient solution converged in {itercount} iterations')
            
            done = True
            force_coeffs = np.loadtxt(f'./{self.casename}/postProcessing/forceCoeffsIncompressible/0/forceCoeffs.dat')
            cl = np.average(force_coeffs[-100:,3])
            cd = np.average(force_coeffs[-100:,2])
                        
            if self.reward_type == 1:
                # lift to drag ratio as the reward
                reward = cl/cd
            elif self.reward_type == 2:
                reward = -cd
        
        # update the state at the end of an episode
        if self.states_type == 1:
            self.state = np.array([self.vx]) 
        elif self.states_type == 2:
            None
        
        # write the data for debugging, post-processing, etc.
        row = np.array([self.worker_index,self.vx,res_ux,res_uy,res_p,
                        res_nutilda,itercount,reward])
    
        row = np.concatenate((row,action_cp))
        
        with open(self.csvfile, 'a') as csvfile:
            np.savetxt(csvfile, np.reshape(row,[1,-1]), delimiter=",")
                                
        return np.array(self.state), reward, done, {}
    
    #function for rendering instantaneous figures, animations, etc. 
    # (retained to keep the code OpenAI Gym optimal)
    def render(self, mode='human'):
        self.n +=0

    # close the training session
    def close(self):
        return 0

#%% 
#-----------------------------------------------------------------------------#
# test the class step function
#-----------------------------------------------------------------------------#

if __name__ == '__main__':
        
    # create the instance of a class    
    env_config = {}
    
    # parameters
    env_config['write_interval'] = 100          # write the OpenFoam solution every 100 iterations
    env_config['end_time'] = 500                # maximum number of iterations of the CFD solver
    env_config['test'] = True                   # flag for testing the class 
    env_config['worker_index'] = 1              # identifier for the processor
    env_config['res_ux_tol'] = 1.0e-4           # residual tolerance for the velocity in x-direction
    env_config['res_uy_tol'] = 1.0e-4           # residual tolerance for the velocity in y-direction
    env_config['res_p_tol'] = 1.0e-4            # residual tolerance for the pressure equation
    env_config['res_nutilda_tol'] = 1.0e-4      # residual tolerance for the nutilda in Spallart-Almaras model
    env_config['reward_type'] = 1               # 1: terminal, 2: at each time step
    env_config['states_type'] = 1               # 1: single state, 2: k states history
    env_config['vx'] = 25.75                    # inlet velocity
    data = np.loadtxt('control_points_range.csv',delimiter=',',skiprows=1, usecols=range(1,4))
    env_config['controlparams_low'] = data[:,1] # lower limit for control parameters
    env_config['controlparams_high'] = data[:,2]# upper limit for control parameters
    
    # create the instance of a class    
    dp = shape_optimization(env_config)   
    
    done = False
    
    # read the action from .csv file
    actions_a1 = np.loadtxt('naca0012_cp_optimized.csv',delimiter=',')

    k = 0    
    start = time.time()
    state_0 = dp.reset()
    reward = 0
    print(k, ' ', state_0, ' ', reward, ' ', done)
    
    # 1 episode = 1 action
    for i in range(1):
        while not done:
            # reset the envrionment
            state_0 = dp.reset()
            
            # action needs to be in list data type
            action = actions_a1.tolist()
            
            # perform an action in the environment
            state, reward, done, dict_empty = dp.step(action)
            k = k + 1
            print(k, ' ', state, ' ', reward, ' ', done)
        
        print('CPU Time = ', time.time() - start)
        done = False